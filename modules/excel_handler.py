"""
Excel Islemleri Modulu
Excel dosyasi okuma, yazma ve format guncelleme islemleri
"""

import pandas as pd
import os
from typing import List, Dict, Optional
from openpyxl import load_workbook
from openpyxl.styles import numbers, Font, Fill, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from field_registry import standard_columns

class ExcelHandler:
    """Excel dosyasi islemleri icin sinif"""
    
    # ⚠️ KRİTİK: Standart sütun sırası - ASLA DEĞİŞTİRME!
    # Bu sıra Excel dosyası formatını belirler. Değiştirirsen:
    # 1. Mevcut Excel dosyaları uyumsuz olur
    # 2. _format_guncelle() fonksiyonunu güncelle
    # 3. gui_widgets.py içindeki Treeview sütunlarını güncelle
    # 4. form_handler.py içindeki form alanlarını güncelle
    # 5. HAND_OFF_DOKUMANTASYON.md dosyasını güncelle
    # Not: Meta kolonlar da bu standarda dahildir.
    STANDART_SUTUN_SIRASI = standard_columns()
    
    def __init__(self, excel_dosyasi: str = None):
        """
        Args:
            excel_dosyasi: Excel dosyasi yolu (None ise masaüstünde oluşturulur)
            
        ⚠️ DİKKAT: Varsayılan olarak masaüstünde "Kutuphanem.xlsx" oluşturulur
        """
        if excel_dosyasi is None:
            # Masaüstü yolunu bul
            desktop_path = os.path.join(os.path.expanduser("~"), "Desktop")
            self.excel_dosyasi = os.path.join(desktop_path, "Kutuphanem.xlsx")
        else:
            self.excel_dosyasi = excel_dosyasi
    
    def yukle(self) -> List[Dict]:
        """
        Excel dosyasindan kitap listesini yukler ve format gunceller
        
        Returns:
            Kitap listesi (dict listesi)
        """
        if not os.path.exists(self.excel_dosyasi):
            return []
        
        try:
            df = pd.read_excel(self.excel_dosyasi, engine='openpyxl')
            df = self._ensure_columns(df)
            
            # Format kontrolü ve güncelleme
            format_guncellenmeli = self._format_kontrol_et(df)
            
            # Verileri yükle
            kitap_listesi = df.to_dict('records')
            
            # NaN değerleri boş string'e çevir
            for kitap in kitap_listesi:
                for key, value in kitap.items():
                    if pd.isna(value):
                        kitap[key] = ""
            
            # Format güncellenmeli ise kaydet
            if format_guncellenmeli and kitap_listesi:
                try:
                    df = self._format_guncelle(df)
                    df.to_excel(self.excel_dosyasi, index=False, engine='openpyxl')
                    # Güncellenmiş verileri tekrar yükle
                    kitap_listesi = df.to_dict('records')
                    for kitap in kitap_listesi:
                        for key, value in kitap.items():
                            if pd.isna(value):
                                kitap[key] = ""
                except Exception as e:
                    # Güncelleme başarısız olursa devam et
                    print(f"Format güncelleme hatası: {e}")
            
            return kitap_listesi
            
        except Exception as e:
            print(f"Excel yükleme hatası: {e}")
            return []
    
    def kaydet(self, kitap_listesi: List[Dict]) -> bool:
        """
        Kitap listesini Excel dosyasına kaydeder
        
        Args:
            kitap_listesi: Kaydedilecek kitap listesi
            
        Returns:
            Başarılı ise True
        """
        try:
            # DataFrame oluştur
            df = pd.DataFrame(kitap_listesi)
            
            # Eksik sütunları ekle
            for sutun in self.STANDART_SUTUN_SIRASI:
                if sutun not in df.columns:
                    df[sutun] = ""
            
            # Sütun sırasını düzenle
            df = df[self.STANDART_SUTUN_SIRASI]
            
            # Yıl sütunlarını sayısal formata çevir
            df = self._yil_sutunlarini_formatla(df)
            
            # Excel'e kaydet
            df.to_excel(self.excel_dosyasi, index=False, engine='openpyxl')
            
            # Excel dosyasını açıp hücre formatlarını ayarla
            self._excel_format_ayarla()
            
            return True
            
        except Exception as e:
            print(f"Excel kaydetme hatası: {e}")
            return False
    
    def sablon_olustur(self, dosya_yolu: str) -> bool:
        """
        Boş Excel şablonu oluşturur (sadece A1: "Kitap", B1: "Yazar")
        
        Args:
            dosya_yolu: Şablonun kaydedileceği yol
            
        Returns:
            Başarılı ise True
            
        ⚠️ DİKKAT: Şablon sadece 2 hücre içerir (A1: "Kitap", B1: "Yazar")
        - Hiçbir formatlama yok (border, bold, vb.)
        - Diğer sütunlar otomatik doldurma ile gelir
        - Şablon formatını değiştirirsen kullanıcıları bilgilendir
        - HAND_OFF_DOKUMANTASYON.md dosyasını güncelle
        """
        try:
            from openpyxl import Workbook
            
            # Yeni workbook oluştur
            wb = Workbook()
            ws = wb.active
            
            # Sadece A1 ve B1 hücrelerine değer yaz (hiçbir formatlama yok)
            ws['A1'] = "Kitap"
            ws['B1'] = "Yazar"
            
            # Dosyayı kaydet
            wb.save(dosya_yolu)
            wb.close()
            
            return True
        except Exception as e:
            print(f"Şablon oluşturma hatası: {e}")
            return False
    
    def disaridan_yukle(self, dosya_yolu: str) -> Optional[List[Dict]]:
        """
        Dışarıdan Excel dosyası yükler ve parse eder
        
        Args:
            dosya_yolu: Yüklenecek Excel dosyası yolu
            
        Returns:
            Kitap listesi veya None (hata durumunda)
        """
        try:
            df = pd.read_excel(dosya_yolu, engine='openpyxl')
            
            # "Kitap" sütununu "Kitap Adı" olarak eşleştir (şablon uyumluluğu için)
            if 'Kitap' in df.columns and 'Kitap Adı' not in df.columns:
                df = df.rename(columns={'Kitap': 'Kitap Adı'})
            
            # Zorunlu sütun kontrolü
            gerekli_sutunlar = ['Kitap Adı', 'Yazar']
            eksik_sutunlar = [sutun for sutun in gerekli_sutunlar if sutun not in df.columns]
            
            if eksik_sutunlar:
                raise ValueError(f"Eksik sütunlar: {', '.join(eksik_sutunlar)}")

            # Zorunlu kolonlar dogrulandiktan sonra eksik meta kolonlarini tamamla
            df = self._ensure_columns(df)
            
            # Boş satırları filtrele
            df = df.dropna(subset=['Kitap Adı', 'Yazar'], how='all')
            df = df[df['Kitap Adı'].notna() & df['Yazar'].notna()]
            
            # Sayısal değerleri güvenli şekilde string'e çevir ve filtrele
            df['Kitap Adı'] = df['Kitap Adı'].fillna('').astype(str).str.strip()
            df['Yazar'] = df['Yazar'].fillna('').astype(str).str.strip()
            df = df[(df['Kitap Adı'] != '') & (df['Yazar'] != '')]
            
            # DataFrame'i sözlük listesine çevir
            kitaplar = df.to_dict('records')
            
            # NaN değerleri boş string'e çevir
            for kitap in kitaplar:
                for key, value in kitap.items():
                    if pd.isna(value):
                        kitap[key] = ""
            
            return kitaplar
            
        except Exception as e:
            print(f"Dışarıdan yükleme hatası: {e}")
            return None
    
    def dosya_acik_mi(self) -> bool:
        """
        Excel dosyasının açık olup olmadığını kontrol eder
        
        Returns:
            Açık ise True
        """
        if not os.path.exists(self.excel_dosyasi):
            return False
        
        try:
            with open(self.excel_dosyasi, 'r+b'):
                pass
            return False
        except PermissionError:
            return True
        except:
            return False
    
    def _format_kontrol_et(self, df: pd.DataFrame) -> bool:
        """
        Excel dosyas?n?n format?n?n g?ncel olup olmad???n? kontrol eder

        Args:
            df: DataFrame

        Returns:
            G?ncellenmeli ise True
        """
        # Eksik s?tun varsa format g?ncellenmeli
        for sutun in self.STANDART_SUTUN_SIRASI:
            if sutun not in df.columns:
                return True

        if len(df.columns) > 1:
            # ?kinci s?tun "Yazar" de?ilse eski format
            if df.columns[1] != "Yazar":
                return True
        elif "Yazar" not in df.columns:
            # Yazar s?tunu yoksa eski format
            return True

        return False

    def _format_guncelle(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        DataFrame'i yeni formata g?re g?nceller

        Args:
            df: G?ncellenecek DataFrame

        Returns:
            G?ncellenmi? DataFrame
        """
        # Kolonlar? standarda uyarla
        df = self._ensure_columns(df)

        # Y?l s?tunlar?n? say?sal formata ?evir
        df = self._yil_sutunlarini_formatla(df)

        return df

    def _ensure_columns(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Eksik kolonlar? ekler ve kolon s?ras?n? standarda g?re d?zenler.
        """
        for sutun in self.STANDART_SUTUN_SIRASI:
            if sutun not in df.columns:
                df[sutun] = ""

        return df[self.STANDART_SUTUN_SIRASI]

    def _yil_sutunlarini_formatla(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Yıl sütunlarını sayısal formata çevirir (tek yıllar sayı, aralıklar metin kalır)
        
        Args:
            df: DataFrame
            
        Returns:
            Formatlanmış DataFrame
        """
        yil_sutunlari = ["İlk Yayınlanma Tarihi", "Anlatı Yılı"]
        
        for sutun in yil_sutunlari:
            if sutun in df.columns:
                # Her değeri kontrol et
                def yil_formatla(deger):
                    if pd.isna(deger) or deger == "":
                        return ""
                    
                    # String'e çevir
                    deger_str = str(deger).strip()
                    
                    # Boş ise boş döndür
                    if not deger_str:
                        return ""
                    
                    # Aralık formatı kontrolü (örn: "1865-1869")
                    if "-" in deger_str:
                        # Aralık formatı - metin olarak bırak
                        return deger_str
                    
                    # Tek yıl formatı - sayısal formata çevir
                    try:
                        yil = int(float(deger_str))
                        # Makul yıl aralığı kontrolü (1000-3000)
                        if 1000 <= yil <= 3000:
                            return yil
                        else:
                            # Makul aralık dışındaysa metin olarak bırak
                            return deger_str
                    except (ValueError, TypeError):
                        # Sayıya çevrilemezse metin olarak bırak
                        return deger_str
                
                # Sütunu formatla
                df[sutun] = df[sutun].apply(yil_formatla)
        
        return df
    
    def _excel_format_ayarla(self):
        """
        Excel dosyasını açıp kütüphane temalı formatlama uygular:
        - Otomatik sütun genişlikleri ve satır yükseklikleri
        - Kılavuz çizgileri kapalı
        - Tüm sayfanın arka plan rengi (boş hücreler dahil)
        - Başlık satırı: koyu kahverengi arka plan, beyaz yazı, bold, 12pt, border
        - Veri satırları: zebra striping (açık bej/açık sarı), normal, 11pt
        - Hizalama: Sola yaslı ve dikey ortalanmış (konusu hariç - wrap text)
        - Georgia font (kütüphane temalı)
        - Sürprizler: Freeze panes, başlık border, otomatik satır yüksekliği
        """
        try:
            wb = load_workbook(self.excel_dosyasi)
            ws = wb.active
            
            # Kılavuz çizgilerini kaldır
            ws.sheet_view.showGridLines = False
            
            # Başlık satırını dondur (scroll yaparken başlık görünür kalır)
            ws.freeze_panes = "A2"
            
            # Kütüphane temalı renkler
            BASLIK_BG = "8B4513"  # Koyu kahverengi
            BASLIK_FG = "FFFFFF"  # Beyaz
            VERI_BG_1 = "F5E6D3"  # Açık bej
            VERI_BG_2 = "FFF8DC"  # Açık sarı
            SAYFA_BG = "FAF5EF"   # Çok açık krem (tüm sayfa arka planı)
            
            # Border stilleri
            thin_border = Side(style='thin', color='8B4513')
            baslik_border = Border(
                left=thin_border,
                right=thin_border,
                top=thin_border,
                bottom=Side(style='medium', color='8B4513')  # Alt border daha kalın
            )
            
            # Başlık satırı formatı
            baslik_satiri = 1
            baslik_font = Font(name="Georgia", size=12, bold=True, color=BASLIK_FG)
            baslik_fill = PatternFill(start_color=BASLIK_BG, end_color=BASLIK_BG, fill_type="solid")
            baslik_alignment = Alignment(horizontal="left", vertical="center")
            
            # Veri satırı fontları
            veri_font = Font(name="Georgia", size=11, bold=False)
            # Konusu için özel hizalama (wrap text ile)
            konusu_alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            # Diğer hücreler için hizalama (sola yaslı, dikey ortalanmış)
            normal_alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
            
            # "Konusu" sütununu bul
            konusu_col_idx = None
            max_col = ws.max_column
            max_row = ws.max_row
            
            # Önce sütun başlıklarını bul ve "Konusu" sütununu tespit et
            for col_idx in range(1, max_col + 1):
                cell = ws.cell(row=baslik_satiri, column=col_idx)
                if cell.value == "Konusu":
                    konusu_col_idx = col_idx
                    break
            
            # Sütun başlıklarını formatla ve sütun genişliklerini ayarla
            sutun_indeksleri = {}
            
            for col_idx in range(1, max_col + 1):
                cell = ws.cell(row=baslik_satiri, column=col_idx)
                sutun_adi = cell.value
                
                # Başlık formatı
                cell.font = baslik_font
                cell.fill = baslik_fill
                cell.alignment = baslik_alignment
                cell.border = baslik_border
                
                # Yıl sütunlarını kaydet
                if sutun_adi in ["İlk Yayınlanma Tarihi", "Anlatı Yılı"]:
                    sutun_indeksleri[sutun_adi] = col_idx
                
                # Sütun genişliğini otomatik ayarla
                col_letter = get_column_letter(col_idx)
                
                # Başlık genişliğini hesapla
                if sutun_adi:
                    baslik_genislik = len(str(sutun_adi)) + 3
                else:
                    baslik_genislik = 10
                
                # Veri genişliğini hesapla (tüm satırları kontrol et)
                veri_genislik = 0
                for row_idx in range(baslik_satiri + 1, max_row + 1):
                    data_cell = ws.cell(row=row_idx, column=col_idx)
                    if data_cell.value:
                        cell_len = len(str(data_cell.value))
                        # Konusu için özel genişlik (daha geniş)
                        if col_idx == konusu_col_idx:
                            if cell_len > 80:
                                cell_len = 80
                        else:
                            # Uzun metinler için maksimum genişlik sınırı
                            if cell_len > 50:
                                cell_len = 50
                        veri_genislik = max(veri_genislik, cell_len)
                
                # Sütun genişliğini ayarla
                if col_idx == konusu_col_idx:
                    # Konusu sütunu daha geniş
                    sutun_genislik = max(25, min(70, max(baslik_genislik, veri_genislik + 3)))
                else:
                    sutun_genislik = max(12, min(60, max(baslik_genislik, veri_genislik + 2)))
                ws.column_dimensions[col_letter].width = sutun_genislik
            
            # Tüm sayfanın arka planını ayarla (boş hücreler dahil)
            # Maksimum 1000 satır ve 50 sütun formatla (performans için)
            max_format_row = min(max_row + 100, 1000)  # Veri + 100 boş satır veya maksimum 1000
            max_format_col = min(max_col, 50)
            
            # Veri satırlarını formatla (zebra striping)
            for row_idx in range(baslik_satiri + 1, max_format_row + 1):
                # Zebra striping: çift satırlar açık bej, tek satırlar açık sarı
                if row_idx % 2 == 0:
                    row_fill = PatternFill(start_color=VERI_BG_1, end_color=VERI_BG_1, fill_type="solid")
                else:
                    row_fill = PatternFill(start_color=VERI_BG_2, end_color=VERI_BG_2, fill_type="solid")
                
                # Boş satırlar için sayfa arka plan rengi
                if row_idx > max_row:
                    row_fill = PatternFill(start_color=SAYFA_BG, end_color=SAYFA_BG, fill_type="solid")
                
                # Satır yüksekliğini hesapla (içeriğe göre)
                max_line_count = 1
                
                for col_idx in range(1, max_format_col + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    
                    # Arka plan rengi
                    if row_idx <= max_row:
                        cell.fill = row_fill
                    else:
                        cell.fill = PatternFill(start_color=SAYFA_BG, end_color=SAYFA_BG, fill_type="solid")
                    
                    # Font ve hizalama
                    if row_idx <= max_row:
                        cell.font = veri_font
                        
                        # Konusu sütunu için özel hizalama
                        if col_idx == konusu_col_idx:
                            cell.alignment = konusu_alignment
                            # Konusu için satır sayısını hesapla (wrap text için)
                            if cell.value:
                                text = str(cell.value)
                                # Yaklaşık satır sayısı hesapla (sütun genişliğine göre)
                                col_width = ws.column_dimensions[get_column_letter(col_idx)].width
                                chars_per_line = int(col_width * 1.2)  # Yaklaşık karakter sayısı
                                if chars_per_line > 0:
                                    line_count = max(1, (len(text) // chars_per_line) + 1)
                                    max_line_count = max(max_line_count, min(line_count, 5))  # Maksimum 5 satır
                        else:
                            cell.alignment = normal_alignment
                    else:
                        # Boş satırlar için varsayılan format
                        cell.font = veri_font
                        cell.alignment = normal_alignment
                    
                    # Yıl sütunlarını sayısal formata çevir
                    if col_idx in sutun_indeksleri.values() and row_idx <= max_row:
                        if cell.value is not None and cell.value != "":
                            deger_str = str(cell.value)
                            if "-" not in deger_str:
                                try:
                                    yil = int(float(deger_str))
                                    if 1000 <= yil <= 3000:
                                        cell.value = yil
                                        cell.number_format = '0'  # Sayısal format (ondalık yok)
                                except (ValueError, TypeError):
                                    pass
                
                # Satır yüksekliğini otomatik ayarla
                if row_idx == baslik_satiri:
                    ws.row_dimensions[row_idx].height = 28  # Başlık satırı daha yüksek
                elif row_idx <= max_row:
                    # İçeriğe göre dinamik yükseklik (minimum 18, maksimum 75)
                    # Her satır için yaklaşık 15pt + (satır sayısı * 12pt)
                    row_height = max(18, min(75, 15 + (max_line_count * 12)))
                    ws.row_dimensions[row_idx].height = row_height
                else:
                    # Boş satırlar için standart yükseklik
                    ws.row_dimensions[row_idx].height = 18
            
            # Boş sütunlar için de arka plan rengi ayarla
            for col_idx in range(max_col + 1, max_format_col + 1):
                for row_idx in range(1, max_format_row + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    if row_idx == baslik_satiri:
                        cell.fill = baslik_fill
                    else:
                        cell.fill = PatternFill(start_color=SAYFA_BG, end_color=SAYFA_BG, fill_type="solid")
            
            wb.save(self.excel_dosyasi)
            wb.close()
            
        except Exception as e:
            # Format ayarlama başarısız olursa devam et (kritik değil)
            print(f"Excel format ayarlama hatası: {e}")
